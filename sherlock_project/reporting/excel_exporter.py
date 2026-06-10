"""
Excel Exporter Module

Exports search results to Excel (.xlsx) format using openpyxl.
"""

import os
from datetime import datetime
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    numbers
)
from openpyxl.utils import get_column_letter

from sherlock_project.result import QueryResult, QueryStatus


class ExcelExporter:
    """Export search results to Excel format"""

    @staticmethod
    def export(
        results: List[QueryResult],
        username: str,
        output_path: Optional[str] = None,
    ) -> str:
        """Export results to Excel file

        Args:
            results: List of QueryResult objects
            username: The username that was searched
            output_path: Optional custom output path

        Returns:
            Path to the generated Excel file
        """
        if output_path is None:
            output_path = f"{username}_report.xlsx"

        wb = Workbook()

        # ===== Summary Sheet =====
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Styles
        header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        title_font = Font(name='Calibri', size=18, bold=True, color='1A237E')
        subtitle_font = Font(name='Calibri', size=11, color='666666')
        stat_label_font = Font(name='Calibri', size=11, bold=True)
        stat_value_font = Font(name='Calibri', size=11)

        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC'),
        )

        # Title
        ws_summary.merge_cells('A1:D1')
        ws_summary['A1'] = 'Sherlock OSINT Report'
        ws_summary['A1'].font = title_font

        ws_summary.merge_cells('A2:D2')
        ws_summary['A2'] = f'Username Analysis: {username}'
        ws_summary['A2'].font = subtitle_font

        ws_summary.merge_cells('A3:D3')
        ws_summary['A3'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws_summary['A3'].font = subtitle_font

        # Summary stats
        total = len(results)
        found = sum(1 for r in results if r.status == QueryStatus.CLAIMED)
        available = sum(1 for r in results if r.status == QueryStatus.AVAILABLE)
        unknown = sum(1 for r in results if r.status in (QueryStatus.UNKNOWN, QueryStatus.ILLEGAL, QueryStatus.WAF))

        ws_summary['A5'] = 'Statistic'
        ws_summary['B5'] = 'Value'
        ws_summary['A5'].font = header_font
        ws_summary['B5'].font = header_font
        ws_summary['A5'].fill = header_fill
        ws_summary['B5'].fill = header_fill
        ws_summary['A5'].alignment = header_alignment
        ws_summary['B5'].alignment = header_alignment

        stats = [
            ('Total Sites Checked', total),
            ('Username Found', found),
            ('Username Not Found', available),
            ('Unknown/Error', unknown),
        ]

        for i, (label, value) in enumerate(stats, start=6):
            ws_summary[f'A{i}'] = label
            ws_summary[f'A{i}'].font = stat_label_font
            ws_summary[f'B{i}'] = value
            ws_summary[f'B{i}'].font = stat_value_font
            ws_summary[f'A{i}'].border = thin_border
            ws_summary[f'B{i}'].border = thin_border

        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 20

        # ===== Detailed Results Sheet =====
        ws_details = wb.create_sheet(title='Detailed Results')

        # Headers
        headers = ['#', 'Site', 'Status', 'Response Time (s)', 'URL']
        detail_header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        detail_header_fill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
        detail_header_align = Alignment(horizontal='center', vertical='center')

        for col, header in enumerate(headers, start=1):
            cell = ws_details.cell(row=1, column=col, value=header)
            cell.font = detail_header_font
            cell.fill = detail_header_fill
            cell.alignment = detail_header_align
            cell.border = thin_border

        # Styles for data rows
        found_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
        found_font = Font(name='Calibri', size=10, color='2E7D32')
        not_found_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
        not_found_font = Font(name='Calibri', size=10, color='C62828')
        default_font = Font(name='Calibri', size=10)

        # Sort: claimed first, then alphabetical
        sorted_results = sorted(
            results,
            key=lambda r: (0 if r.status == QueryStatus.CLAIMED else 1, r.site_name.lower())
        )

        for idx, result in enumerate(sorted_results, start=1):
            row = idx + 1

            status_text = {
                QueryStatus.CLAIMED: "Found",
                QueryStatus.AVAILABLE: "Not Found",
                QueryStatus.UNKNOWN: "Unknown",
                QueryStatus.ILLEGAL: "Illegal",
                QueryStatus.WAF: "WAF Blocked",
            }.get(result.status, str(result.status))

            time_text = result.query_time if result.query_time else ""

            # Write data
            ws_details.cell(row=row, column=1, value=idx).font = default_font
            ws_details.cell(row=row, column=1).alignment = Alignment(horizontal='center')

            site_cell = ws_details.cell(row=row, column=2, value=result.site_name)
            site_cell.font = default_font

            status_cell = ws_details.cell(row=row, column=3, value=status_text)
            status_cell.alignment = Alignment(horizontal='center')
            if result.status == QueryStatus.CLAIMED:
                status_cell.font = found_font
                status_cell.fill = found_fill
            elif result.status == QueryStatus.AVAILABLE:
                status_cell.font = not_found_font
                status_cell.fill = not_found_fill
            else:
                status_cell.font = default_font

            time_cell = ws_details.cell(row=row, column=4, value=time_text)
            time_cell.font = default_font
            time_cell.alignment = Alignment(horizontal='center')
            if isinstance(time_text, (int, float)):
                time_cell.number_format = '0.00'

            url_cell = ws_details.cell(row=row, column=5, value=result.site_url_user or "N/A")
            url_cell.font = default_font

            # Apply borders
            for col in range(1, 6):
                ws_details.cell(row=row, column=col).border = thin_border

        # Column widths
        ws_details.column_dimensions['A'].width = 6
        ws_details.column_dimensions['B'].width = 40
        ws_details.column_dimensions['C'].width = 16
        ws_details.column_dimensions['D'].width = 18
        ws_details.column_dimensions['E'].width = 80

        # Freeze header row
        ws_details.freeze_panes = 'A2'

        # Auto-filter
        ws_details.auto_filter.ref = f'A1:E{len(sorted_results) + 1}'

        # Save
        wb.save(output_path)
        return output_path


def export_excel(
    results: List[QueryResult],
    username: str,
    output_path: Optional[str] = None,
) -> str:
    """Convenience function to export results to Excel"""
    return ExcelExporter.export(results, username, output_path)